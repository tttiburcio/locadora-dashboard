import openpyxl

path = r'c:\Users\ADM\Documents\locadora-dashboard\Locadora.xlsx'
wb = openpyxl.load_workbook(path)
sheet_name = next(s for s in wb.sheetnames if 'MANUTENCOES' in s)
ws = wb[sheet_name]

headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
col_id = headers.index('IDOrdServ') + 1
col_pos = next(i for i, h in enumerate(headers) if h and 'Posi' in str(h) and 'Pneu' in str(h)) + 1
col_desc = next(i for i, h in enumerate(headers) if h and 'Descri' in str(h)) + 1

count = 0
for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=col_id).value == 'OS-2025-0112':
        old_desc = ws.cell(row=r, column=col_desc).value
        ws.cell(row=r, column=col_pos).value = "DIANTEIRO"
        if old_desc and "TRASEIRO" in str(old_desc).upper():
             new_d = str(old_desc).replace("TRASEIROS", "DIANTEIROS").replace("TRASEIRO", "DIANTEIRO")
             ws.cell(row=r, column=col_desc).value = new_d
        count += 1

if count > 0:
    wb.save(path)
    print(f"Excel revert SUCCESS. Updated {count} rows.")
else:
    print("No rows found to revert.")
wb.close()
