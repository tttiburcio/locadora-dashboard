import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
wb = openpyxl.load_workbook('Locadora.xlsx', read_only=True)

for name in wb.sheetnames:
    if 'MANUT' in name.upper():
        print(f'Sheet: {name!r}')
        ws = wb[name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print('Headers:', headers)
        # Print rows for DYG5G03
        row_count = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if any(v for v in row if v is not None):
                # Find placa/IDVeiculo column
                print(f'  Row {i}: {row[:15]}')  # show first 15 cols
                row_count += 1
            if row_count > 40:
                print('  ...(truncated)')
                break
wb.close()
