"""
sync_km_dyg5g03.py
Sincroniza a quilometragem real da placa DYG5G03 a partir da API de rastreamento (tracker_status.db)
para os registros de manutenção no banco locadora.db e na planilha Locadora.xlsx.

Estratégia:
  1. Reconstrói o odômetro diário do DYG5G03 via tracker_status.db (mesma lógica da api/main.py)
  2. Para cada OS/manutenção, busca o km acumulado na data de execução
  3. Atualiza manutencoes.km e ordens_servico.km no locadora.db
  4. Atualiza a coluna KM na aba MANUTENCOES do Locadora.xlsx
"""

import sys
import io
import os
import sqlite3
import pandas as pd
import openpyxl
from datetime import date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

PLACA = "DYG5G03"
TRACKER_DB = r"C:\Users\vinic\Documents\mapWS\dados\tracker_status.db"
LOCADORA_DB = r"C:\Users\vinic\Documents\clone\locadora.db"
LOCADORA_XL = r"C:\Users\vinic\Documents\clone\Locadora.xlsx"


# ── 1. Reconstruir odômetro diário ────────────────────────────────────────────

def build_odometer(placa: str) -> pd.DataFrame:
    """Retorna DataFrame com colunas [Data (date), KM_Acumulado (float)] para a placa."""
    conn = sqlite3.connect(TRACKER_DB)
    df_base = pd.read_sql(
        "SELECT KM_Base, Data_Base FROM status_atual WHERE Placa = ?",
        conn, params=(placa,)
    )
    df_hist = pd.read_sql(
        "SELECT Data, KM FROM historico_rastreamento WHERE Placa = ? ORDER BY Data ASC",
        conn, params=(placa,)
    )
    conn.close()

    if df_base.empty or df_hist.empty:
        raise RuntimeError(f"Sem dados no tracker para {placa}")

    base_km = float(df_base["KM_Base"].iloc[0])
    base_date = pd.to_datetime(df_base["Data_Base"].iloc[0], dayfirst=True)

    df_hist["Data"] = pd.to_datetime(df_hist["Data"], errors="coerce")
    df_hist = df_hist.dropna(subset=["Data"]).sort_values("Data")
    df_hist["KM_Relative"] = df_hist["KM"].cumsum()

    try:
        relative_at_base = df_hist[df_hist["Data"] <= base_date]["KM_Relative"].iloc[-1]
    except IndexError:
        relative_at_base = df_hist["KM_Relative"].iloc[0]

    shift = base_km - relative_at_base
    df_hist["KM_Acumulado"] = df_hist["KM_Relative"] + shift

    # Preenche lacunas (forward-fill por data)
    min_date = df_hist["Data"].min()
    max_date = df_hist["Data"].max()
    full_range = pd.date_range(start=min_date, end=max_date, freq="D")
    df_full = pd.DataFrame({"Data": full_range})
    df_full = df_full.merge(df_hist[["Data", "KM_Acumulado"]], on="Data", how="left")
    df_full["KM_Acumulado"] = df_full["KM_Acumulado"].ffill().bfill()
    df_full["Data"] = df_full["Data"].dt.date
    return df_full.set_index("Data")


def km_at_date(odometer: pd.DataFrame, ref_date) -> int | None:
    """Retorna o odômetro arredondado para a data fornecida (ou anterior mais próxima)."""
    if ref_date is None:
        return None
    if isinstance(ref_date, str):
        ref_date = pd.to_datetime(ref_date).date()
    elif hasattr(ref_date, "date"):
        ref_date = ref_date.date()

    if ref_date in odometer.index:
        return round(float(odometer.at[ref_date, "KM_Acumulado"]))

    # Busca a data anterior mais próxima
    earlier = odometer[odometer.index <= ref_date]
    if not earlier.empty:
        return round(float(earlier["KM_Acumulado"].iloc[-1]))
    return None


# ── 2. Atualizar banco locadora.db ────────────────────────────────────────────

def update_db(odometer: pd.DataFrame):
    conn = sqlite3.connect(LOCADORA_DB)
    cur = conn.cursor()

    cur.execute("SELECT id FROM frota WHERE placa = ?", (PLACA,))
    row = cur.fetchone()
    if not row:
        raise RuntimeError(f"Placa {PLACA} não encontrada em frota")
    id_veiculo = row[0]

    # --- manutencoes ---
    cur.execute(
        "SELECT id, data_execucao, data_entrada, km FROM manutencoes WHERE id_veiculo = ?",
        (id_veiculo,)
    )
    manut_rows = cur.fetchall()
    manut_updated = 0
    print(f"\n=== manutencoes ({len(manut_rows)} registros) ===")
    for mid, data_exec, data_entrada, km_atual in manut_rows:
        ref = data_exec or data_entrada
        km_novo = km_at_date(odometer, ref)
        status = "ATUALIZADO" if km_novo != km_atual else "igual"
        print(f"  id={mid:4d} | data={ref} | km_atual={km_atual} -> km_novo={km_novo} | {status}")
        if km_novo is not None and km_novo != km_atual:
            cur.execute("UPDATE manutencoes SET km = ? WHERE id = ?", (km_novo, mid))
            manut_updated += 1

    # --- ordens_servico ---
    cur.execute(
        "SELECT id, numero_os, data_execucao, data_entrada, km FROM ordens_servico WHERE id_veiculo = ?",
        (id_veiculo,)
    )
    os_rows = cur.fetchall()
    os_updated = 0
    print(f"\n=== ordens_servico ({len(os_rows)} registros) ===")
    for oid, numero_os, data_exec, data_entrada, km_atual in os_rows:
        ref = data_exec or data_entrada
        km_novo = km_at_date(odometer, ref)
        status = "ATUALIZADO" if km_novo != km_atual else "igual"
        print(f"  id={oid:4d} | {numero_os} | data={ref} | km_atual={km_atual} -> km_novo={km_novo} | {status}")
        if km_novo is not None and km_novo != km_atual:
            cur.execute("UPDATE ordens_servico SET km = ? WHERE id = ?", (km_novo, oid))
            os_updated += 1

    conn.commit()
    conn.close()
    print(f"\nDB: {manut_updated} manutencoes + {os_updated} ordens_servico atualizadas")


# ── 3. Atualizar Locadora.xlsx ────────────────────────────────────────────────

def update_excel(odometer: pd.DataFrame):
    wb = openpyxl.load_workbook(LOCADORA_XL)
    sheet_name = next((s for s in wb.sheetnames if "MANUT" in s.upper()), None)
    if not sheet_name:
        raise RuntimeError("Aba MANUTENCOES não encontrada no Excel")

    ws = wb[sheet_name]

    # Detectar índices das colunas-chave pelo cabeçalho (row 1)
    header = {cell.value: cell.column for cell in ws[1] if cell.value}
    col_placa = header.get("Placa")
    col_km = header.get("KM")
    col_data_exec = header.get("DataExecução")
    col_id_ord = header.get("IDOrdServ")

    if not all([col_placa, col_km, col_data_exec]):
        raise RuntimeError(f"Colunas não encontradas. Header: {list(header.keys())}")

    print(f"\n=== Excel: aba '{sheet_name}' ===")
    print(f"  col_placa={col_placa}, col_km={col_km}, col_data_exec={col_data_exec}, col_id_ord={col_id_ord}")

    # Busca datas de execução por IDOrdServ no DB (fallback quando Excel não tiver a data)
    conn = sqlite3.connect(LOCADORA_DB)
    os_dates = {}
    for numero_os, data_exec, data_entrada in conn.execute(
        "SELECT numero_os, data_execucao, data_entrada FROM ordens_servico WHERE id_veiculo = "
        "(SELECT id FROM frota WHERE placa = ?)", (PLACA,)
    ):
        os_dates[numero_os] = data_exec or data_entrada
    conn.close()

    excel_updated = 0
    excel_skipped = 0

    for row in ws.iter_rows(min_row=2):
        placa_cell = row[col_placa - 1]
        if placa_cell.value != PLACA:
            continue

        km_cell = row[col_km - 1]
        data_exec_cell = row[col_data_exec - 1]
        id_ord_cell = row[col_id_ord - 1] if col_id_ord else None

        # Determina data de referência
        ref_date = None
        if data_exec_cell.value and not isinstance(data_exec_cell.value, str):
            ref_date = data_exec_cell.value  # datetime do openpyxl
        elif id_ord_cell and id_ord_cell.value in os_dates:
            ref_date = os_dates[id_ord_cell.value]

        km_novo = km_at_date(odometer, ref_date)
        row_num = km_cell.row
        km_atual = km_cell.value

        if km_novo is not None:
            print(f"  row={row_num:4d} | {id_ord_cell.value if id_ord_cell else '?'} | data={ref_date} | km: {km_atual} -> {km_novo}")
            km_cell.value = km_novo
            excel_updated += 1
        else:
            print(f"  row={row_num:4d} | {id_ord_cell.value if id_ord_cell else '?'} | SEM DATA — ignorado")
            excel_skipped += 1

    wb.save(LOCADORA_XL)
    wb.close()
    print(f"\nExcel: {excel_updated} células atualizadas, {excel_skipped} ignoradas")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"Reconstruindo odômetro do {PLACA} via tracker_status.db...")
    odometer = build_odometer(PLACA)
    print(f"  Período: {odometer.index.min()} a {odometer.index.max()}")
    print(f"  KM range: {odometer['KM_Acumulado'].min():.0f} ~ {odometer['KM_Acumulado'].max():.0f} km")

    update_db(odometer)
    update_excel(odometer)

    print("\n✓ Sincronização concluída.")
